import tkinter as tk
from tkinter import messagebox, ttk
import webbrowser
import urllib.parse
import threading
import time
import json
import re
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from datetime import datetime
import os

class NaverSearchCompleteApp:
    def __init__(self, root):
        self.root = root
        self.root.title("네이버 검색 프로그램 (엑셀 저장 지원)")
        self.root.geometry("1400x900")
        
        # 키워드 히스토리 파일
        self.history_file = "search_history.json"
        self.search_history = self.load_search_history()
        
        # GUI 구성
        self.setup_gui()
        
    def load_search_history(self):
        """검색 히스토리 로드"""
        try:
            if os.path.exists(self.history_file):
                with open(self.history_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except:
            pass
        return []
        
    def save_search_history(self):
        """검색 히스토리 저장"""
        try:
            with open(self.history_file, 'w', encoding='utf-8') as f:
                json.dump(self.search_history[-20:], f, ensure_ascii=False, indent=2)
        except:
            pass
            
    def add_to_history(self, keyword):
        """검색 히스토리에 키워드 추가"""
        if keyword and keyword.strip():
            keyword = keyword.strip()
            if keyword in self.search_history:
                self.search_history.remove(keyword)
            self.search_history.append(keyword)
            self.update_combobox_values()
            self.save_search_history()
            
    def update_combobox_values(self):
        """첫 번째 Combobox 값들 업데이트 (최신순)"""
        if self.keyword_entries and hasattr(self.keyword_entries[0], 'config'):
            # 첫 번째가 Combobox인 경우만
            if hasattr(self.keyword_entries[0], '__setitem__'):  # Combobox 확인
                self.update_combobox_values_for_combo(self.keyword_entries[0])
        
    def on_combobox_click(self, event=None):
        """Combobox 클릭시 히스토리 업데이트"""
        self.update_combobox_values()
        
    def add_keyword_entry(self):
        """키워드 입력폼 추가"""
        row = len(self.keyword_entries)
        
        # 키워드 프레임
        keyword_frame = ttk.Frame(self.keywords_container)
        keyword_frame.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
        
        # 라벨
        if row == 0:
            label_text = "기본 키워드:"
        else:
            label_text = f"키워드 {row + 1}:"
        
        label = ttk.Label(keyword_frame, text=label_text, width=12)
        label.grid(row=0, column=0, padx=(0, 10))
        
        # 자동완성 기능을 위한 Combobox (첫 번째만)
        if row == 0:
            keyword_combo = ttk.Combobox(keyword_frame, width=32)
            keyword_combo.bind('<Button-1>', self.on_combobox_click)
            keyword_combo.bind('<FocusIn>', self.on_combobox_click)
            keyword_combo.bind('<Delete>', self.delete_selected_history)  # Del 키로 삭제
            keyword_combo.bind('<Control-d>', self.delete_selected_history)  # Ctrl+D로 삭제
            self.update_combobox_values_for_combo(keyword_combo)
            
            # 히스토리 삭제 버튼 추가
            history_delete_btn = ttk.Button(keyword_frame, text="❌", width=3,
                                          command=lambda: self.delete_selected_history_from_menu(keyword_combo))
            history_delete_btn.grid(row=0, column=2, padx=(5, 10))
            self.history_delete_button = history_delete_btn
        else:
            keyword_combo = ttk.Entry(keyword_frame, width=35)
        
        keyword_combo.grid(row=0, column=1, padx=(0, 5))
        keyword_combo.bind('<Return>', lambda e: self.search_keyword())
        
        # AND 표시 (첫 번째가 아닌 경우)
        if row > 0:
            and_label = ttk.Label(keyword_frame, text="AND", font=('맑은 고딕', 9, 'bold'))
            and_label.grid(row=0, column=2, padx=(10, 10))
        
        self.keyword_entries.append(keyword_combo)
        self.keyword_frames.append(keyword_frame)
        
        # 제거 버튼 상태 업데이트
        self.update_remove_button_state()
        
    def show_history_menu(self, event):
        """검색 히스토리 우클릭 메뉴 표시"""
        try:
            combo = event.widget
            if not hasattr(combo, 'current'):  # Combobox인지 확인
                return
                
            # 컨텍스트 메뉴 생성
            context_menu = tk.Menu(self.root, tearoff=0)
            context_menu.add_command(label="🗑️ 선택된 항목 삭제", command=lambda: self.delete_selected_history_from_menu(combo))
            context_menu.add_separator()
            context_menu.add_command(label="🗑️ 모든 히스토리 삭제", command=self.clear_all_history)
            context_menu.add_separator()
            context_menu.add_command(label="💡 팁: Del키로 선택 항목 삭제", state='disabled')
            
            # 메뉴 표시
            context_menu.tk_popup(event.x_root, event.y_root)
            
        except Exception as e:
            print(f"히스토리 메뉴 오류: {e}")
            
    def delete_selected_history(self, event):
        """선택된 히스토리 항목 삭제 (키보드 단축키)"""
        try:
            combo = event.widget
            if not hasattr(combo, 'current'):  # Combobox인지 확인
                return
                
            self.delete_selected_history_from_menu(combo)
            
        except Exception as e:
            print(f"히스토리 삭제 오류: {e}")
            
    def delete_selected_history_from_menu(self, combo):
        """선택된 히스토리 항목 삭제"""
        try:
            current_text = combo.get().strip()
            if not current_text:
                messagebox.showwarning("알림", "삭제할 키워드를 선택하거나 입력해주세요.")
                return
                
            if current_text in self.search_history:
                # 확인 대화상자
                if messagebox.askyesno("삭제 확인", f"'{current_text}' 키워드를 히스토리에서 삭제하시겠습니까?"):
                    self.search_history.remove(current_text)
                    self.save_search_history()
                    
                    # Combobox 업데이트
                    self.update_combobox_values_for_combo(combo)
                    combo.set("")  # 입력 필드 비우기
                    
                    messagebox.showinfo("완료", f"'{current_text}' 키워드가 삭제되었습니다.")
            else:
                messagebox.showwarning("알림", f"'{current_text}' 키워드가 히스토리에 없습니다.")
                
        except Exception as e:
            messagebox.showerror("오류", f"히스토리 삭제 중 오류가 발생했습니다:\n{str(e)}")
            
    def clear_all_history(self):
        """모든 검색 히스토리 삭제"""
        try:
            if not self.search_history:
                messagebox.showinfo("알림", "삭제할 히스토리가 없습니다.")
                return
                
            # 확인 대화상자
            if messagebox.askyesno("삭제 확인", f"모든 검색 히스토리({len(self.search_history)}개)를 삭제하시겠습니까?\n\n이 작업은 되돌릴 수 없습니다."):
                self.search_history.clear()
                self.save_search_history()
                
                # 첫 번째 Combobox 업데이트
                if self.keyword_entries and hasattr(self.keyword_entries[0], 'config'):
                    self.update_combobox_values_for_combo(self.keyword_entries[0])
                    self.keyword_entries[0].set("")
                    
                messagebox.showinfo("완료", "모든 검색 히스토리가 삭제되었습니다.")
                
        except Exception as e:
            messagebox.showerror("오류", f"히스토리 삭제 중 오류가 발생했습니다:\n{str(e)}")

    def add_exclude_keyword_entry(self):
        """제외 키워드 입력폼 추가"""
        if not hasattr(self, 'exclude_keyword_entries'):
            self.exclude_keyword_entries = []
        if not hasattr(self, 'exclude_keyword_frames'):
            self.exclude_keyword_frames = []
            
        row = len(self.exclude_keyword_entries)
        
        # 제외 키워드 프레임
        exclude_frame = ttk.Frame(self.exclude_keywords_container)
        exclude_frame.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
        
        # 라벨
        label_text = f"제외 키워드 {row + 1}:"
        label = ttk.Label(exclude_frame, text=label_text, width=15)
        label.grid(row=0, column=0, padx=(0, 10))
        
        # 제외 키워드 입력폼
        exclude_entry = ttk.Entry(exclude_frame, width=30)
        exclude_entry.grid(row=0, column=1, padx=(0, 10))
        exclude_entry.bind('<KeyRelease>', lambda e: self.apply_exclude_filters())
        exclude_entry.bind('<Return>', lambda e: self.apply_exclude_filters())
        
        # 제거 버튼
        remove_btn = ttk.Button(exclude_frame, text="❌", width=3,
                               command=lambda idx=row: self.remove_exclude_keyword_entry(idx))
        remove_btn.grid(row=0, column=2, padx=(0, 10))
        
        self.exclude_keyword_entries.append(exclude_entry)
        self.exclude_keyword_frames.append(exclude_frame)
        
    def remove_exclude_keyword_entry(self, idx):
        """특정 제외 키워드 입력폼 제거"""
        if idx < len(self.exclude_keyword_entries):
            # 프레임 제거
            self.exclude_keyword_frames[idx].destroy()
            
            # 리스트에서 제거
            del self.exclude_keyword_entries[idx]
            del self.exclude_keyword_frames[idx]
            
            # 나머지 프레임들 재배치
            for i, frame in enumerate(self.exclude_keyword_frames):
                frame.grid(row=i, column=0, sticky=(tk.W, tk.E), pady=2)
                
            # 필터 재적용
            self.apply_exclude_filters()

    def remove_keyword_entry(self):
        """마지막 키워드 입력폼 제거"""
        if len(self.keyword_entries) > 1:
            # 마지막 프레임과 입력폼 제거
            last_frame = self.keyword_frames.pop()
            last_entry = self.keyword_entries.pop()
            
            last_frame.destroy()
            
            # 제거 버튼 상태 업데이트
            self.update_remove_button_state()
            
    def update_remove_button_state(self):
        """제거 버튼 활성화/비활성화 상태 업데이트"""
        if hasattr(self, 'remove_keyword_button'):  # 버튼이 존재하는지 확인
            if len(self.keyword_entries) <= 1:
                self.remove_keyword_button.config(state='disabled')
            else:
                self.remove_keyword_button.config(state='normal')
            
    def update_combobox_values_for_combo(self, combo):
        """특정 Combobox에 대해 값들 업데이트"""
        reversed_history = list(reversed(self.search_history))
        combo['values'] = reversed_history
        
    def apply_exclude_filters(self):
        """제외 키워드 필터 적용"""
        if not hasattr(self, 'original_search_results') or not self.original_search_results:
            return
            
        # 제외할 키워드들 수집
        exclude_keywords = []
        if hasattr(self, 'exclude_keyword_entries'):
            for entry in self.exclude_keyword_entries:
                keyword = entry.get().strip().lower()
                if keyword:
                    exclude_keywords.append(keyword)
        
        # 필터링 적용
        filtered_results = []
        for result in self.original_search_results:
            should_exclude = False
            
            # 제외 키워드가 제목이나 내용에 포함되어 있는지 확인
            for exclude_keyword in exclude_keywords:
                if (exclude_keyword in result['title'].lower() or 
                    exclude_keyword in result['content'].lower()):
                    should_exclude = True
                    break
            
            if not should_exclude:
                filtered_results.append(result)
        
        # 필터링된 결과로 업데이트
        self.search_results = filtered_results
        self.refresh_treeview()
        
        # 상태 업데이트
        excluded_count = len(self.original_search_results) - len(filtered_results)
        if excluded_count > 0:
            exclude_keywords_text = ", ".join(exclude_keywords) if exclude_keywords else ""
            self.update_status(f"검색 완료 - 총 {len(self.original_search_results)}개 결과 중 {excluded_count}개 제외 ('{exclude_keywords_text}'), {len(filtered_results)}개 표시")
        else:
            self.update_status(f"검색 완료 - 총 {len(filtered_results)}개 결과")
            
    def enable_exclude_buttons(self):
        """검색 완료 후 제외 버튼들 활성화"""
        if hasattr(self, 'exclude_buttons'):
            for btn in self.exclude_buttons:
                btn.config(state='normal')
        
    def setup_gui(self):
        """GUI 구성"""
        # 메뉴바 생성
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        
        # 파일 메뉴
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="파일", menu=file_menu)
        file_menu.add_command(label="엑셀 파일 불러오기", command=self.load_excel_file)
        file_menu.add_separator()
        file_menu.add_command(label="엑셀로 저장", command=self.save_to_excel)
        
        # 보기 메뉴
        view_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="보기", menu=view_menu)
        view_menu.add_command(label="전체 결과", command=self.show_all_results)
        view_menu.add_command(label="즐겨찾기만", command=self.show_favorites_only_toggle)
        view_menu.add_separator()
        view_menu.add_command(label="즐겨찾기 초기화", command=self.clear_favorites)
        
        # 메인 프레임
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 검색 입력 프레임
        search_frame = ttk.LabelFrame(main_frame, text="검색 키워드 (AND 조건)", padding="10")
        search_frame.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # 키워드 입력폼들을 관리할 리스트
        self.keyword_entries = []
        self.keyword_frames = []
        
        # 키워드 컨테이너 프레임
        self.keywords_container = ttk.Frame(search_frame)
        self.keywords_container.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # 첫 번째 키워드 입력폼 추가
        self.add_keyword_entry()
        
        # 버튼 프레임
        button_frame = ttk.Frame(search_frame)
        button_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E))
        
        # 키워드 추가/제거 버튼
        self.add_keyword_button = ttk.Button(button_frame, text="+ 키워드 추가", command=self.add_keyword_entry)
        self.add_keyword_button.pack(side=tk.LEFT, padx=(0, 5))
        
        self.remove_keyword_button = ttk.Button(button_frame, text="- 키워드 제거", command=self.remove_keyword_entry)
        self.remove_keyword_button.pack(side=tk.LEFT, padx=(0, 10))
        
        # 검색 버튼
        self.search_button = ttk.Button(button_frame, text="🔍 검색", command=self.search_keyword)
        self.search_button.pack(side=tk.LEFT, padx=(10, 0))
        
        # 제외 키워드 프레임 추가
        exclude_frame = ttk.LabelFrame(main_frame, text="제외 키워드 (결과에서 제외)", padding="10")
        exclude_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # 제외 키워드 입력폼들을 관리할 리스트 초기화
        self.exclude_keyword_entries = []
        self.exclude_keyword_frames = []
        
        # 제외 키워드 컨테이너 프레임
        self.exclude_keywords_container = ttk.Frame(exclude_frame)
        self.exclude_keywords_container.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # 제외 키워드 버튼 프레임
        exclude_button_frame = ttk.Frame(exclude_frame)
        exclude_button_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E))
        
        # 제외 키워드 추가 버튼
        self.add_exclude_keyword_button = ttk.Button(exclude_button_frame, text="+ 제외 키워드 추가", command=self.add_exclude_keyword_entry)
        self.add_exclude_keyword_button.pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Label(exclude_button_frame, text="💡 키워드 입력 후 자동으로 결과에서 제외됩니다", 
                 font=('맑은 고딕', 9), foreground='gray').pack(side=tk.LEFT, padx=(10, 0))
        
        # 옵션 프레임
        option_frame = ttk.Frame(main_frame)
        option_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.open_browser_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(option_frame, text="브라우저에서 검색 결과 열기", 
                       variable=self.open_browser_var).grid(row=0, column=0)
        
        # 진행 상태
        self.progress = ttk.Progressbar(main_frame, mode='indeterminate')
        self.progress.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.status_label = ttk.Label(main_frame, text="준비 (엑셀 저장 지원)")
        self.status_label.grid(row=4, column=0, columnspan=3, sticky=(tk.W,), pady=(0, 10))
        
        # 검색 결과 리스트 (전체 영역 사용)
        self.setup_result_list(main_frame)
        
        # 그리드 설정
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(5, weight=1)  # 결과 리스트가 row 5로 이동
        
    def setup_result_list(self, parent):
        """검색 결과 리스트 구성 (내용 요약 포함)"""
        result_frame = ttk.LabelFrame(parent, text="검색 결과", padding="10")
        result_frame.grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 트리뷰 설정 (내용 요약 컬럼 추가, 폰트 크기 및 행 높이 증가)
        style = ttk.Style()
        style.configure("Treeview", font=('맑은 고딕', 11), rowheight=30)  # 폰트 크기 및 행 높이 증가
        style.configure("Treeview.Heading", font=('맑은 고딕', 12, 'bold'))  # 헤딩 폰트도 증가
        
        self.tree = ttk.Treeview(result_frame, columns=('favorite', 'type', 'content', 'date'), show='tree headings', height=20)
        scrollbar_v = ttk.Scrollbar(result_frame, orient=tk.VERTICAL, command=self.tree.yview)
        scrollbar_h = ttk.Scrollbar(result_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=scrollbar_v.set, xscrollcommand=scrollbar_h.set)
        
        # 컬럼 설정 (정렬 기능 포함)
        self.tree.heading('#0', text='제목', command=lambda: self.sort_column('#0', False))
        self.tree.heading('favorite', text='⭐', command=lambda: self.sort_column('favorite', False))
        self.tree.heading('type', text='유형', command=lambda: self.sort_column('type', False))
        self.tree.heading('content', text='내용 요약', command=lambda: self.sort_column('content', False))
        self.tree.heading('date', text='날짜 ⬇️', command=lambda: self.sort_column('date', False))  # 기본 내림차순 표시
        
        self.tree.column('#0', width=300, minwidth=200)
        self.tree.column('favorite', width=50, minwidth=40)
        self.tree.column('type', width=80, minwidth=60)
        self.tree.column('content', width=480, minwidth=300)
        self.tree.column('date', width=150, minwidth=100)
        
        # 정렬 상태 추적
        self.sort_column_name = 'date'
        self.sort_reverse = True  # 기본적으로 최신순 (내림차순)
        
        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar_v.grid(row=0, column=1, sticky=(tk.N, tk.S))
        scrollbar_h.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        result_frame.columnconfigure(0, weight=1)
        result_frame.rowconfigure(0, weight=1)
        
        # 버튼 영역
        button_frame = ttk.Frame(result_frame)
        button_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 0))
        
        self.open_link_button = ttk.Button(button_frame, text="브라우저에서 열기", 
                                         command=self.open_selected_link, state='disabled')
        self.open_link_button.pack(side=tk.LEFT, padx=(0, 10))
        
        self.excel_save_button = ttk.Button(button_frame, text="📊 엑셀 저장", 
                                          command=self.save_to_excel, state='disabled')
        self.excel_save_button.pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Label(button_frame, text="💡 팁: 더블클릭으로 링크 열기").pack(side=tk.LEFT)
        
        # 이벤트 바인딩
        self.tree.bind('<ButtonRelease-1>', self.on_item_click)
        self.tree.bind('<Double-1>', self.on_item_double_click)
        self.tree.bind('<Button-1>', self.on_favorite_click)  # 즐겨찾기 클릭 처리
        
        self.search_results = []
        self.favorites = set()  # 즐겨찾기한 항목들의 링크를 저장
        self.show_favorites_only = False  # 즐겨찾기만 보기 모드
        
    def parse_date_for_sorting(self, date_str):
        """날짜 문자열을 정렬용 datetime 객체로 변환"""
        from datetime import datetime, timedelta
        import re
        
        if not date_str or date_str == "날짜 정보 없음":
            return datetime.min
        
        current_time = datetime.now()
        
        # 상대적 시간 표현 처리
        relative_patterns = [
            (r'(\d+)분\s*전', 'minutes'),
            (r'(\d+)시간\s*전', 'hours'),
            (r'(\d+)일\s*전', 'days'),
            (r'(\d+)개월\s*전', 'months'),
            (r'(\d+)년\s*전', 'years')
        ]
        
        for pattern, unit in relative_patterns:
            match = re.search(pattern, date_str)
            if match:
                value = int(match.group(1))
                if unit == 'minutes':
                    return current_time - timedelta(minutes=value)
                elif unit == 'hours':
                    return current_time - timedelta(hours=value)
                elif unit == 'days':
                    return current_time - timedelta(days=value)
                elif unit == 'months':
                    return current_time - timedelta(days=value * 30)  # 근사치
                elif unit == 'years':
                    return current_time - timedelta(days=value * 365)  # 근사치
        
        # 절대적 날짜 표현 처리
        absolute_patterns = [
            r'(\d{4})\.(\d{1,2})\.(\d{1,2})',  # 2024.01.15
            r'(\d{1,2})\.(\d{1,2})\.(\d{4})',  # 15.01.2024
            r'(\d{1,2})\.(\d{1,2})\.',         # 01.15. (현재 년도)
        ]
        
        for pattern in absolute_patterns:
            match = re.search(pattern, date_str)
            if match:
                try:
                    if len(match.groups()) == 3:
                        if len(match.group(1)) == 4:  # YYYY.MM.DD
                            year, month, day = int(match.group(1)), int(match.group(2)), int(match.group(3))
                        else:  # DD.MM.YYYY
                            day, month, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
                    else:  # MM.DD. (현재 년도)
                        month, day = int(match.group(1)), int(match.group(2))
                        year = current_time.year
                    
                    return datetime(year, month, day)
                except:
                    continue
        
        # 파싱 실패시 현재 시간 반환
        return current_time
        
    def is_within_week(self, date_str):
        """날짜가 1달 내인지 확인 (최근 게시물 필터링)"""
        from datetime import datetime, timedelta
        import re
        
        if not date_str or date_str == "날짜 정보 없음":
            return False
        
        current_time = datetime.now()
        one_month_ago = current_time - timedelta(days=30)  # 1달로 변경
        
        # 상대적 시간 표현 처리 (한국어 + 영어)
        relative_patterns = [
            (r'(\d+)분\s*전', 'minutes'),
            (r'(\d+)시간\s*전', 'hours'),
            (r'(\d+)일\s*전', 'days'),
            (r'(\d+)개월\s*전', 'months'),
            (r'(\d+)년\s*전', 'years'),
            # 영어 패턴 추가
            (r'(\d+)\s*minute[s]?\s*ago', 'minutes'),
            (r'(\d+)\s*hour[s]?\s*ago', 'hours'), 
            (r'(\d+)\s*day[s]?\s*ago', 'days'),
            (r'(\d+)\s*week[s]?\s*ago', 'weeks'),
            (r'(\d+)\s*month[s]?\s*ago', 'months'),
            (r'(\d+)\s*year[s]?\s*ago', 'years')
        ]
        
        for pattern, unit in relative_patterns:
            match = re.search(pattern, date_str)
            if match:
                value = int(match.group(1))
                if unit == 'minutes' or unit == 'hours':
                    return True  # 몇 분/시간 전은 확실히 1달 내
                elif unit == 'days':
                    return value <= 30  # 30일 이내
                elif unit == 'weeks':
                    return value <= 4  # 4주 이내
                elif unit == 'months':
                    return value <= 1  # 1개월 이내
                elif unit == 'years':
                    return False  # 몇 년 전은 1달 밖
        
        # 절대적 날짜 표현 처리
        absolute_patterns = [
            r'(\d{4})\.(\d{1,2})\.(\d{1,2})',  # 2024.01.15
            r'(\d{1,2})\.(\d{1,2})\.(\d{4})',  # 15.01.2024
            r'(\d{1,2})\.(\d{1,2})\.',         # 01.15. (현재 년도)
        ]
        
        for pattern in absolute_patterns:
            match = re.search(pattern, date_str)
            if match:
                try:
                    if len(match.groups()) == 3:
                        if len(match.group(1)) == 4:  # YYYY.MM.DD
                            year, month, day = int(match.group(1)), int(match.group(2)), int(match.group(3))
                        else:  # DD.MM.YYYY
                            day, month, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
                    else:  # MM.DD. (현재 년도)
                        month, day = int(match.group(1)), int(match.group(2))
                        year = current_time.year
                    
                    post_date = datetime(year, month, day)
                    return post_date >= one_month_ago
                except:
                    continue
        
        # 파싱 실패시 포함 (안전하게)
        return True
        
    def sort_column(self, col, reverse):
        """컬럼 정렬"""
        # 정렬 상태 업데이트
        if self.sort_column_name == col:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_column_name = col
            self.sort_reverse = reverse
        
        # 헤딩 화살표 업데이트
        for column in ['#0', 'favorite', 'type', 'content', 'date']:
            if column == col:
                arrow = ' ⬇️' if self.sort_reverse else ' ⬆️'
                if column == '#0':
                    self.tree.heading(column, text=f'제목{arrow}')
                elif column == 'favorite':
                    self.tree.heading(column, text=f'⭐{arrow}')
                elif column == 'type':
                    self.tree.heading(column, text=f'유형{arrow}')
                elif column == 'content':
                    self.tree.heading(column, text=f'내용 요약{arrow}')
                elif column == 'date':
                    self.tree.heading(column, text=f'날짜{arrow}')
            else:
                # 다른 컬럼들은 화살표 제거
                if column == '#0':
                    self.tree.heading(column, text='제목')
                elif column == 'favorite':
                    self.tree.heading(column, text='⭐')
                elif column == 'type':
                    self.tree.heading(column, text='유형')
                elif column == 'content':
                    self.tree.heading(column, text='내용 요약')
                elif column == 'date':
                    self.tree.heading(column, text='날짜')
        
        # 검색 결과 정렬
        if col == 'date':
            # 날짜는 특별 처리
            self.search_results.sort(key=lambda x: self.parse_date_for_sorting(x['date']), 
                                   reverse=self.sort_reverse)
        elif col == '#0':
            # 제목 정렬
            self.search_results.sort(key=lambda x: x['title'], reverse=self.sort_reverse)
        elif col == 'favorite':
            # 즐겨찾기 정렬 (즐겨찾기한 것이 먼저)
            self.search_results.sort(key=lambda x: x['link'] in self.favorites, reverse=self.sort_reverse)
        elif col == 'type':
            # 유형 정렬
            self.search_results.sort(key=lambda x: x['type'], reverse=self.sort_reverse)
        elif col == 'content':
            # 내용 정렬
            self.search_results.sort(key=lambda x: x['content'], reverse=self.sort_reverse)
        
        # 트리뷰 업데이트
        self.refresh_treeview()
        
    def refresh_treeview(self):
        """트리뷰 내용 새로고침"""
        # 기존 항목들 삭제
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # 표시할 결과 필터링
        display_results = self.search_results
        if self.show_favorites_only:
            display_results = [result for result in self.search_results if result['link'] in self.favorites]
        
        # 정렬된 결과로 다시 채우기
        for i, result in enumerate(display_results):
            content_summary = result['content'][:100] + "..." if len(result['content']) > 100 else result['content']
            favorite_icon = "⭐" if result['link'] in self.favorites else "☆"
            
            self.tree.insert('', 'end', 
                           text=result['title'],
                           values=(favorite_icon, result['type'], content_summary, result['date']),
                           tags=(str(i),))
        

        
    def update_status(self, message):
        self.status_label.config(text=message)
        self.root.update_idletasks()
        
    def search_keyword(self):
        # 모든 키워드 입력폼에서 값 수집
        keywords = []
        for entry in self.keyword_entries:
            keyword = entry.get().strip()
            if keyword:
                keywords.append(keyword)
        
        if not keywords:
            messagebox.showwarning("경고", "최소 하나의 검색 키워드를 입력해주세요.")
            return
            
        # 첫 번째 키워드를 히스토리에 추가
        self.add_to_history(keywords[0])
        
        # AND 조건으로 키워드 조합
        combined_keyword = " ".join(keywords)  # 간단한 조합 (공백으로 구분)
        display_keyword = " AND ".join(keywords)  # 표시용
            
        # 검색 결과 초기화
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.search_results = []
        self.current_keyword = display_keyword  # 현재 키워드 저장 (표시용)
        self.search_keywords = keywords  # 개별 키워드들 저장
        self.open_link_button.config(state='disabled')
        self.excel_save_button.config(state='disabled')
            
        self.search_button.config(state='disabled')
        self.progress.start()
        
        thread = threading.Thread(target=self.perform_search, args=(combined_keyword,))
        thread.daemon = True
        thread.start()
        
    def perform_search(self, keyword):
        encoded_keyword = urllib.parse.quote(keyword)
        
        # 브라우저에서 결과 열기
        if self.open_browser_var.get():
            cafe_url = f"https://search.naver.com/search.naver?cafe_where=&prdtype=0&query={encoded_keyword}&sm=mtb_opt&ssc=tab.cafe.all&st=date&stnm=date&opt_tab=0&nso=so%3Add%2Cp%3Aall"
            blog_url = f"https://search.naver.com/search.naver?ssc=tab.blog.all&sm=tab_jum&query={encoded_keyword}&nso=so%3Add%2Cp%3Aall"
            
            webbrowser.open(cafe_url)
            webbrowser.open(blog_url)
        
        all_results = []
        
        # 카페 검색
        self.root.after(0, self.update_status, "카페 검색 중...")
        cafe_results = self.scrape_naver(encoded_keyword, "cafe")
        all_results.extend(cafe_results)
        
        # 블로그 검색
        self.root.after(0, self.update_status, "블로그 검색 중...")
        blog_results = self.scrape_naver(encoded_keyword, "blog")
        all_results.extend(blog_results)
        
        # 유튜브 검색
        self.root.after(0, self.update_status, "유튜브 검색 중...")
        youtube_results = self.scrape_naver(encoded_keyword, "youtube")
        all_results.extend(youtube_results)
        
        # 모든 결과를 한 번에 업데이트
        self.root.after(0, self.update_all_results, all_results)
        self.root.after(0, self.update_status, f"검색 완료 - 카페 {len(cafe_results)}개, 블로그 {len(blog_results)}개, 유튜브 {len(youtube_results)}개")
        self.root.after(0, self.search_complete)
        
    def scrape_naver(self, encoded_keyword, search_type):
        """네이버 검색 스크래핑 (빠른 버전)"""
        driver = None
        results = []
        
        try:
            print(f"\n=== {search_type} 검색 시작 (빠른 모드) ===")
            
            # Chrome 옵션 (검색 타입별 최적화)
            chrome_options = Options()
            chrome_options.add_argument("--headless")  # 헤드리스 모드로 빠르게
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--disable-extensions")
            chrome_options.add_argument("--disable-images")  # 이미지 로딩 안함
            
            # 유튜브는 JavaScript 필요, 네이버는 비활성화 가능
            if search_type != "youtube":
                chrome_options.add_argument("--disable-javascript")  # JS 비활성화로 빠르게
            
            chrome_options.add_argument("--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36")
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=chrome_options)
            driver.set_page_load_timeout(10)  # 타임아웃 줄임
            
            # URL 설정 (최신순 정렬 포함)
            if search_type == "cafe":
                url = f"https://search.naver.com/search.naver?ssc=tab.cafe.all&query={encoded_keyword}&nso=so%3Add%2Cp%3Aall"
                target_domain = "cafe.naver.com"
            elif search_type == "blog":
                url = f"https://search.naver.com/search.naver?ssc=tab.blog.all&query={encoded_keyword}&nso=so%3Add%2Cp%3Aall"
                target_domain = "blog.naver.com"
            else:  # youtube
                url = f"https://www.youtube.com/results?search_query={encoded_keyword}&sp=CAI%253D"
                target_domain = "youtube.com"
            
            driver.get(url)
            time.sleep(3)  # 유튜브는 조금 더 기다림
            
            if search_type == "youtube":
                # 유튜브 전용 처리
                return self.scrape_youtube(driver, encoded_keyword)
            
            # 네이버 검색 처리 (카페, 블로그)
            # 링크 수집
            all_links = driver.find_elements(By.TAG_NAME, "a")
            print(f"총 {len(all_links)}개 링크 발견")
            
            valid_links = []
            
            for link in all_links:
                try:
                    href = link.get_attribute('href')
                    text = link.text.strip()
                    
                    if (href and target_domain in href and 
                        text and len(text) >= 5 and len(text) <= 200 and
                        not any(skip in text.lower() for skip in 
                               ['더보기', 'more', '카페', '블로그', '검색', '네이버', '이전', '다음', '전체', '설정', '로그인'])):
                        
                        # 한글 포함 확인
                        if any(ord(char) >= 0xAC00 and ord(char) <= 0xD7A3 for char in text):
                            
                            content = self.extract_content(link)
                            date = self.extract_date(link)
                            
                            # 날짜 필터링: 1달 내 데이터만
                            if not self.is_within_week(date):
                                print(f"❌ {search_type} 날짜 필터링: {text[:30]}... (날짜: {date})")
                                continue
                            
                            valid_links.append({
                                'type': search_type,
                                'title': text,
                                'link': href,
                                'content': content if content else f"[{search_type} 게시글] 내용을 가져올 수 없습니다.",
                                'date': date if date else "날짜 정보 없음"
                            })
                            
                            print(f"✅ {search_type} 결과: {text[:50]}... (날짜: {date})")
                            
                            if len(valid_links) >= 20:
                                break
                                
                except:
                    continue
            
            results = valid_links
            print(f"{search_type} 총 {len(results)}개 결과 추출")
            
        except Exception as e:
            print(f"{search_type} 검색 오류: {e}")
            
        finally:
            if driver:
                try:
                    driver.quit()
                except:
                    pass
                time.sleep(2)
            
        return results
        
    def scrape_youtube(self, driver, encoded_keyword):
        """유튜브 검색 스크래핑"""
        results = []
        try:
            # 페이지 스크롤해서 더 많은 결과 로드
            driver.execute_script("window.scrollTo(0, 1000);")
            time.sleep(2)
            
            # 유튜브 비디오 요소들 찾기
            video_elements = driver.find_elements(By.CSS_SELECTOR, 'div[id="dismissible"]')
            print(f"유튜브에서 {len(video_elements)}개 비디오 발견")
            
            for video in video_elements:
                try:
                    # 제목과 링크 추출
                    title_element = video.find_element(By.CSS_SELECTOR, 'a[id="video-title"]')
                    title = title_element.get_attribute('title') or title_element.text.strip()
                    video_url = title_element.get_attribute('href')
                    
                    if not title or not video_url or len(title) < 5:
                        continue
                    
                    # 전체 URL로 변환
                    if video_url.startswith('/watch'):
                        video_url = f"https://www.youtube.com{video_url}"
                    
                    # 채널명 추출
                    try:
                        channel_element = video.find_element(By.CSS_SELECTOR, 'a[class*="yt-simple-endpoint"][href*="/@"], a[class*="yt-simple-endpoint"][href*="/channel/"], a[class*="yt-simple-endpoint"][href*="/c/"]')
                        channel_name = channel_element.text.strip()
                    except:
                        channel_name = "채널 정보 없음"
                    
                    # 업로드 날짜 추출
                    try:
                        date_elements = video.find_elements(By.CSS_SELECTOR, 'span[class*="style-scope ytd-video-meta-block"], div[id="metadata-line"] span')
                        upload_date = "날짜 정보 없음"
                        for elem in date_elements:
                            text = elem.text.strip()
                            if any(keyword in text for keyword in ['전', '일', '주', '개월', '년', 'ago', 'day', 'week', 'month', 'year']) and any(char.isdigit() for char in text):
                                upload_date = text
                                break
                    except:
                        upload_date = "날짜 정보 없음"
                    
                    # 날짜 필터링
                    if not self.is_within_week(upload_date):
                        print(f"❌ youtube 날짜 필터링: {title[:30]}... (날짜: {upload_date})")
                        continue
                    
                    # 조회수나 기타 정보 추출 (선택사항)
                    try:
                        meta_elements = video.find_elements(By.CSS_SELECTOR, 'span[class*="style-scope ytd-video-meta-block"]')
                        view_count = ""
                        for elem in meta_elements:
                            text = elem.text.strip()
                            if '조회수' in text or 'views' in text.lower():
                                view_count = text
                                break
                    except:
                        view_count = ""
                    
                    # 내용 구성 (채널명 + 조회수)
                    content_parts = [f"채널: {channel_name}"]
                    if view_count:
                        content_parts.append(view_count)
                    content = " | ".join(content_parts)
                    
                    results.append({
                        'type': 'youtube',
                        'title': title,
                        'link': video_url,
                        'content': content,
                        'date': upload_date
                    })
                    
                    print(f"✅ youtube 결과: {title[:50]}... (날짜: {upload_date})")
                    
                    if len(results) >= 20:
                        break
                        
                except Exception as e:
                    continue
            
            print(f"youtube 총 {len(results)}개 결과 추출")
            
        except Exception as e:
            print(f"youtube 검색 오류: {e}")
        
        return results
        
    def extract_content(self, element):
        """내용 추출"""
        try:
            current = element
            for _ in range(6):
                try:
                    current = current.find_element(By.XPATH, "..")
                    elements = current.find_elements(By.CSS_SELECTOR, 
                        "p, div, span, .desc, .description, .content, .summary, .dsc, .txt")
                    
                    for elem in elements:
                        text = elem.text.strip()
                        if (text and len(text) > 30 and text != element.text and 
                            not text.startswith('http') and
                            any(ord(char) >= 0xAC00 and ord(char) <= 0xD7A3 for char in text)):
                            return text[:500]
                except:
                    break
        except:
            pass
        return ""
        
    def extract_date(self, element):
        """날짜 추출 (개선된 버전)"""
        try:
            current = element
            for level in range(8):  # 더 넓은 범위에서 탐색
                try:
                    current = current.find_element(By.XPATH, "..")
                    
                    # 더 다양한 날짜 관련 선택자
                    elements = current.find_elements(By.CSS_SELECTOR, 
                        ".date, .time, .ago, .sub_time, .posting_date, .created_time, "
                        "[class*='date'], [class*='time'], [class*='ago'], "
                        ".txt_inline, .sub_txt, .txt_num, .source_txt")
                    
                    for elem in elements:
                        text = elem.text.strip()
                        if text and 3 <= len(text) <= 30:
                            # 더 정확한 날짜 패턴 매칭
                            date_patterns = [
                                r'\d{4}\.\d{1,2}\.\d{1,2}',  # 2024.01.15
                                r'\d{1,2}\.\d{1,2}\.',        # 01.15.
                                r'\d+일\s*전',                # 3일 전
                                r'\d+시간\s*전',              # 5시간 전
                                r'\d+분\s*전',                # 30분 전
                                r'\d+개월\s*전',              # 2개월 전
                                r'\d+년\s*전'                 # 1년 전
                            ]
                            
                            import re
                            for pattern in date_patterns:
                                if re.search(pattern, text):
                                    return text
                            
                            # 기본 키워드 매칭 (백업용)
                            if any(keyword in text for keyword in 
                                  ['전', '시간', '분', '일', '월', '년', ':', '.']):
                                # 숫자가 포함된 경우만
                                if any(char.isdigit() for char in text):
                                    return text
                
                    # 텍스트에서 직접 날짜 패턴 찾기
                    full_text = current.text
                    if full_text:
                        import re
                        # 상대적 시간 표현 찾기
                        relative_patterns = [
                            r'\d+분\s*전', r'\d+시간\s*전', r'\d+일\s*전',
                            r'\d+개월\s*전', r'\d+년\s*전'
                        ]
                        for pattern in relative_patterns:
                            match = re.search(pattern, full_text)
                            if match:
                                return match.group()
                        
                        # 절대 날짜 표현 찾기
                        absolute_patterns = [
                            r'\d{4}\.\d{1,2}\.\d{1,2}',
                            r'\d{1,2}\.\d{1,2}\.'
                        ]
                        for pattern in absolute_patterns:
                            match = re.search(pattern, full_text)
                            if match:
                                return match.group()
                                
                except:
                    continue
                    
        except:
            pass
        
        # 모든 방법이 실패한 경우 현재 시간 기준으로 표시
        from datetime import datetime
        return datetime.now().strftime("%Y.%m.%d")
        
    def update_all_results(self, all_results):
        """모든 검색 결과를 한 번에 업데이트"""
        # 원본 결과 저장 (필터링용)
        self.original_search_results = all_results.copy()
        self.search_results = all_results
        
        # 날짜순으로 정렬 (최신순)
        if self.search_results:
            self.search_results.sort(key=lambda x: self.parse_date_for_sorting(x['date']), reverse=True)
            
        # 트리뷰 업데이트
        self.refresh_treeview()
        
        # 제외 버튼들 활성화
        if all_results:
            self.enable_exclude_buttons()
        
        # 결과가 없으면 알림
        if not all_results:
            messagebox.showinfo("알림", "검색 결과가 없습니다.")
            
    def update_results_partial(self, new_results):
        """중간 결과 업데이트 (사용 안함 - 호환성 유지)"""
        pass
            
    def update_results_final(self, all_results):
        """최종 결과 업데이트 (사용 안함 - 호환성 유지)"""
        pass
            
    def on_favorite_click(self, event):
        """즐겨찾기 클릭 처리"""
        item = self.tree.identify_row(event.y)
        if item:
            column = self.tree.identify_column(event.x)
            if column == '#1':  # 즐겨찾기 컬럼 (첫 번째 values 컬럼)
                tags = self.tree.item(item, 'tags')
                if tags:
                    # 현재 표시된 결과에서의 인덱스
                    display_index = int(tags[0])
                    
                    # 표시 중인 결과 목록 가져오기
                    display_results = self.search_results
                    if self.show_favorites_only:
                        display_results = [result for result in self.search_results if result['link'] in self.favorites]
                    
                    if 0 <= display_index < len(display_results):
                        result = display_results[display_index]
                        link = result['link']
                        
                        # 즐겨찾기 토글
                        if link in self.favorites:
                            self.favorites.remove(link)
                        else:
                            self.favorites.add(link)
                        
                        # 트리뷰 즉시 업데이트
                        favorite_icon = "⭐" if link in self.favorites else "☆"
                        current_values = list(self.tree.item(item, 'values'))
                        current_values[0] = favorite_icon
                        self.tree.item(item, values=current_values)
                        
                        return  # 다른 이벤트 처리 방지

    def on_item_click(self, event):
        """항목 클릭시 선택 상태 업데이트"""
        selection = self.tree.selection()
        if selection:
            tags = self.tree.item(selection[0], 'tags')
            if tags:
                # 현재 표시된 결과에서의 인덱스
                display_index = int(tags[0])
                
                # 표시 중인 결과 목록 가져오기
                display_results = self.search_results
                if self.show_favorites_only:
                    display_results = [result for result in self.search_results if result['link'] in self.favorites]
                
                if 0 <= display_index < len(display_results):
                    self.selected_result = display_results[display_index]
                    self.open_link_button.config(state='normal')
                    

            
    def on_item_double_click(self, event):
        """더블클릭시 링크 열기"""
        self.open_selected_link()
        
    def show_all_results(self):
        """전체 결과 보기"""
        self.show_favorites_only = False
        self.refresh_treeview()
        
    def show_favorites_only_toggle(self):
        """즐겨찾기만 보기 토글"""
        self.show_favorites_only = True
        self.refresh_treeview()
        
        if not any(result['link'] in self.favorites for result in self.search_results):
            messagebox.showinfo("알림", "즐겨찾기한 항목이 없습니다.")
            
    def clear_favorites(self):
        """즐겨찾기 초기화"""
        if self.favorites:
            if messagebox.askyesno("확인", "모든 즐겨찾기를 삭제하시겠습니까?"):
                self.favorites.clear()
                self.refresh_treeview()
                messagebox.showinfo("완료", "즐겨찾기가 초기화되었습니다.")
        else:
            messagebox.showinfo("알림", "즐겨찾기한 항목이 없습니다.")

    def open_selected_link(self):
        """선택된 링크 열기"""
        if hasattr(self, 'selected_result') and self.selected_result:
            webbrowser.open(self.selected_result['link'])
                
    def search_complete(self):
        self.progress.stop()
        self.search_button.config(state='normal')
        
        # 검색 결과가 있으면 엑셀 저장 버튼 활성화
        if self.search_results:
            self.excel_save_button.config(state='normal')
            
    def save_to_excel(self):
        """검색 결과를 엑셀 파일로 저장"""
        if not self.search_results:
            messagebox.showwarning("경고", "저장할 검색 결과가 없습니다.")
            return
            
        try:
            # 파일 경로 설정
            current_date = datetime.now().strftime("%Y%m%d")
            filename = f"{current_date}-{self.current_keyword}.xlsx"
            filepath = os.path.join(os.path.expanduser("~/Desktop/image-preview-app"), filename)
            
            # 디렉토리가 없으면 생성
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            
            # 시트명 설정 (추출날짜,키워드)
            sheet_name = f"{current_date},{self.current_keyword}"
            # 엑셀 시트명 길이 제한 (31자) 처리
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:28] + "..."
            
            # DataFrame 생성
            data = []
            for result in self.search_results:
                data.append({
                    '즐겨찾기': '⭐' if result['link'] in self.favorites else '☆',
                    '제목': result['title'],
                    '유형': result['type'],
                    '내용요약': result['content'],
                    '날짜': result['date'],
                    '링크': result['link']
                })
            
            df = pd.DataFrame(data)
            
            # 기존 파일이 있는지 확인
            if os.path.exists(filepath):
                # 기존 파일에 시트 추가
                book = load_workbook(filepath)
                
                # 같은 이름의 시트가 있으면 삭제
                if sheet_name in book.sheetnames:
                    book.remove(book[sheet_name])
                
                # 새 시트 생성
                book.create_sheet(sheet_name)
                sheet = book[sheet_name]
                
                # 데이터 추가
                for r in dataframe_to_rows(df, index=False, header=True):
                    sheet.append(r)
                
                # 컬럼 너비 자동 조정
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
                
                book.save(filepath)
                
            else:
                # 새 파일 생성
                with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # 컬럼 너비 자동 조정
                    worksheet = writer.sheets[sheet_name]
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # 저장된 결과 개수
            result_count = len(self.search_results)
            cafe_count = len([r for r in self.search_results if r['type'] == 'cafe'])
            blog_count = len([r for r in self.search_results if r['type'] == 'blog'])
            youtube_count = len([r for r in self.search_results if r['type'] == 'youtube'])
            favorite_count = len([r for r in self.search_results if r['link'] in self.favorites])
            
            messagebox.showinfo("저장 완료", 
                               f"엑셀 파일로 저장되었습니다!\n\n"
                               f"파일: {filename}\n"
                               f"시트: {sheet_name}\n"
                               f"결과: 총 {result_count}개 (카페 {cafe_count}개, 블로그 {blog_count}개, 유튜브 {youtube_count}개)\n"
                               f"즐겨찾기: {favorite_count}개\n"
                               f"위치: ~/Desktop/image-preview-app/")
            
        except Exception as e:
            messagebox.showerror("저장 오류", f"엑셀 저장 중 오류가 발생했습니다:\n{str(e)}")

    def load_excel_file(self):
        """엑셀 파일을 불러와서 리스트에 표시"""
        try:
            from tkinter import filedialog
            
            # 파일 선택 다이얼로그
            file_path = filedialog.askopenfilename(
                title="엑셀 파일 선택",
                filetypes=[
                    ("Excel files", "*.xlsx *.xls"),
                    ("All files", "*.*")
                ],
                initialdir=os.path.expanduser("~/Desktop/image-preview-app")
            )
            
            if not file_path:
                return  # 사용자가 취소한 경우
                
            # 엑셀 파일 읽기
            try:
                # 첫 번째 시트 읽기 (기본)
                df = pd.read_excel(file_path, sheet_name=0)
            except Exception as e:
                # 시트가 여러 개인 경우 시트 선택 다이얼로그
                try:
                    excel_file = pd.ExcelFile(file_path)
                    sheet_names = excel_file.sheet_names
                    
                    if len(sheet_names) > 1:
                        # 간단한 시트 선택 다이얼로그
                        from tkinter import simpledialog
                        sheet_choices = "\n".join([f"{i+1}. {name}" for i, name in enumerate(sheet_names)])
                        choice = simpledialog.askstring(
                            "시트 선택", 
                            f"여러 시트가 있습니다. 번호를 입력하세요:\n\n{sheet_choices}"
                        )
                        
                        if choice and choice.isdigit():
                            sheet_index = int(choice) - 1
                            if 0 <= sheet_index < len(sheet_names):
                                df = pd.read_excel(file_path, sheet_name=sheet_names[sheet_index])
                            else:
                                df = pd.read_excel(file_path, sheet_name=0)
                        else:
                            df = pd.read_excel(file_path, sheet_name=0)
                    else:
                        df = pd.read_excel(file_path, sheet_name=0)
                except Exception as e2:
                    messagebox.showerror("파일 읽기 오류", f"엑셀 파일을 읽을 수 없습니다:\n{str(e2)}")
                    return
            
            # 데이터프레임이 비어있는지 확인
            if df.empty:
                messagebox.showwarning("경고", "선택한 엑셀 파일이 비어있습니다.")
                return
                
            # 컬럼명 정리 (다양한 형식 지원)
            df.columns = df.columns.str.strip()
            
            # 필요한 컬럼들 매핑 (유연한 매핑)
            column_mapping = {
                '제목': ['제목', 'title', 'Title', '타이틀'],
                '유형': ['유형', 'type', 'Type', '타입', '분류'],
                '내용요약': ['내용요약', 'content', 'Content', '내용', '요약', 'summary'],
                '날짜': ['날짜', 'date', 'Date', '작성일', '등록일'],
                '링크': ['링크', 'link', 'Link', 'url', 'URL', '주소'],
                '즐겨찾기': ['즐겨찾기', 'favorite', 'Favorite', '별표']
            }
            
            # 실제 컬럼명 찾기
            actual_columns = {}
            for target_col, possible_names in column_mapping.items():
                for col_name in df.columns:
                    if col_name in possible_names:
                        actual_columns[target_col] = col_name
                        break
                        
            # 필수 컬럼 확인 (제목, 링크)
            if '제목' not in actual_columns or '링크' not in actual_columns:
                missing_cols = []
                if '제목' not in actual_columns:
                    missing_cols.append('제목 (또는 title)')
                if '링크' not in actual_columns:
                    missing_cols.append('링크 (또는 link, url)')
                    
                messagebox.showerror("컬럼 오류", 
                                   f"필수 컬럼이 없습니다:\n{', '.join(missing_cols)}\n\n"
                                   f"현재 컬럼: {', '.join(df.columns.tolist())}")
                return
                
            # 데이터 변환
            loaded_results = []
            favorites_from_file = set()
            
            for _, row in df.iterrows():
                try:
                    # 기본 정보 추출
                    title = str(row[actual_columns['제목']]).strip()
                    link = str(row[actual_columns['링크']]).strip()
                    
                    # 빈 값이면 건너뛰기
                    if not title or title == 'nan' or not link or link == 'nan':
                        continue
                        
                    # 선택적 정보 추출 (없으면 기본값)
                    result_type = str(row[actual_columns.get('유형', '')]).strip() if '유형' in actual_columns else 'excel'
                    if result_type == 'nan':
                        result_type = 'excel'
                        
                    content = str(row[actual_columns.get('내용요약', '')]).strip() if '내용요약' in actual_columns else '엑셀에서 불러온 데이터'
                    if content == 'nan':
                        content = '엑셀에서 불러온 데이터'
                        
                    date = str(row[actual_columns.get('날짜', '')]).strip() if '날짜' in actual_columns else '날짜 정보 없음'
                    if date == 'nan':
                        date = '날짜 정보 없음'
                        
                    # 즐겨찾기 정보
                    if '즐겨찾기' in actual_columns:
                        favorite_value = str(row[actual_columns['즐겨찾기']]).strip()
                        if favorite_value == '⭐' or favorite_value.lower() in ['true', '1', 'yes']:
                            favorites_from_file.add(link)
                    
                    loaded_results.append({
                        'title': title,
                        'type': result_type,
                        'content': content,
                        'date': date,
                        'link': link
                    })
                    
                except Exception as e:
                    continue  # 개별 행 처리 오류는 건너뛰기
                    
            if not loaded_results:
                messagebox.showwarning("경고", "유효한 데이터를 찾을 수 없습니다.")
                return
                
            # 불러온 데이터를 어떻게 처리할지 사용자에게 묻기
            from tkinter import messagebox
            choice = messagebox.askyesnocancel(
                "데이터 처리 방식", 
                f"{len(loaded_results)}개의 데이터를 불러왔습니다.\n\n"
                f"'예': 기존 결과에 추가\n"
                f"'아니오': 기존 결과를 대체\n"
                f"'취소': 작업 취소"
            )
            
            if choice is None:  # 취소
                return
            elif choice:  # 예 - 추가
                # 기존 결과에 추가
                self.search_results.extend(loaded_results)
                self.favorites.update(favorites_from_file)
                mode_text = "추가"
            else:  # 아니오 - 대체
                # 기존 결과 대체
                self.search_results = loaded_results
                self.favorites = favorites_from_file
                mode_text = "대체"
                
            # 원본 결과도 업데이트 (필터링 기능을 위해)
            self.original_search_results = self.search_results.copy()
            
            # 현재 키워드 정보 업데이트
            filename = os.path.basename(file_path)
            self.current_keyword = f"엑셀파일-{filename}"
            
            # 날짜순 정렬
            self.search_results.sort(key=lambda x: self.parse_date_for_sorting(x['date']), reverse=True)
            
            # 트리뷰 업데이트
            self.refresh_treeview()
            
            # 버튼 상태 업데이트
            if self.search_results:
                self.excel_save_button.config(state='normal')
                self.enable_exclude_buttons()
                
            # 상태 업데이트
            total_count = len(self.search_results)
            cafe_count = len([r for r in self.search_results if r['type'] == 'cafe'])
            blog_count = len([r for r in self.search_results if r['type'] == 'blog'])
            youtube_count = len([r for r in self.search_results if r['type'] == 'youtube'])
            excel_count = len([r for r in self.search_results if r['type'] == 'excel'])
            favorite_count = len(favorites_from_file)
            
            result_summary = []
            if cafe_count > 0:
                result_summary.append(f"카페 {cafe_count}개")
            if blog_count > 0:
                result_summary.append(f"블로그 {blog_count}개")
            if youtube_count > 0:
                result_summary.append(f"유튜브 {youtube_count}개")
            if excel_count > 0:
                result_summary.append(f"엑셀 {excel_count}개")
                
            self.update_status(f"엑셀 파일 불러오기 완료 ({mode_text}) - 총 {total_count}개 ({', '.join(result_summary)})")
            
            messagebox.showinfo("불러오기 완료", 
                               f"엑셀 파일을 성공적으로 불러왔습니다!\n\n"
                               f"파일: {filename}\n"
                               f"처리: {mode_text}\n"
                               f"결과: 총 {total_count}개\n"
                               f"즐겨찾기: {favorite_count}개")
                               
        except Exception as e:
            messagebox.showerror("불러오기 오류", f"엑셀 파일 불러오기 중 오류가 발생했습니다:\n{str(e)}")

def main():
    root = tk.Tk()
    root.title("보니의 웹 크롤링 프로그램")
    app = NaverSearchCompleteApp(root)
    
    def on_closing():
        if messagebox.askokcancel("종료", "프로그램을 종료하시겠습니까?"):
            root.destroy()
    
    root.protocol("WM_DELETE_WINDOW", on_closing)
    root.mainloop()

if __name__ == "__main__":
    main() 